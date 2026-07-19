from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import urllib.parse
import urllib.request
from datetime import datetime, timedelta
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

import pdfplumber
import yaml
from openpyxl import load_workbook


MAX_DOWNLOAD_BYTES = 100 * 1024 * 1024
MATERIAL_ID_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_-]{0,127}$")
PROHIBITED_SESSION_FIELDS = {
    "account",
    "authorization",
    "cookie",
    "cookies",
    "password",
    "session",
    "session_id",
    "token",
    "username",
}
LOCATOR_PRIORITY = {
    "ISSUER": 0,
    "REGULATOR_EXCHANGE": 1,
    "NAMED_INSTITUTION": 2,
    "STABLE_WEB": 3,
    "OTHER": 4,
}
ALLOWED_PARSE_STATUSES = {"PARSED", "PARTIAL", "UNSUPPORTED"}
ALLOWED_SEARCH_STATUSES = {"IN_PROGRESS", "SATURATED"}
ACQUISITION_STATUS = "ACQUIRED_UNASSESSED"


def sha256_bytes(content: bytes) -> str:
    return f"sha256:{hashlib.sha256(content).hexdigest()}"


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def parse_timestamp(value: str, field_name: str) -> datetime:
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError as error:
        raise ValueError(f"{field_name} must be an ISO 8601 timestamp: {value}") from error
    if parsed.tzinfo is None:
        raise ValueError(f"{field_name} must include a timezone: {value}")
    return parsed


def write_yaml(path: Path, value: Any) -> None:
    path.write_text(
        yaml.safe_dump(value, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )


def write_json(path: Path, value: Any) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def require_fields(item: dict[str, Any], fields: tuple[str, ...]) -> None:
    missing = [field for field in fields if not item.get(field)]
    if missing:
        raise ValueError(f"material is missing required fields: {', '.join(missing)}")


def validate_material_id(material_id: str, seen_material_ids: set[str]) -> None:
    if not MATERIAL_ID_PATTERN.fullmatch(material_id):
        raise ValueError(
            "material_id must contain only letters, digits, underscores, or hyphens "
            "and must not exceed 128 characters"
        )
    if material_id in seen_material_ids:
        raise ValueError(f"duplicate material_id: {material_id}")
    seen_material_ids.add(material_id)


def find_session_fields(value: Any) -> set[str]:
    prohibited: set[str] = set()
    if isinstance(value, dict):
        for key, nested_value in value.items():
            normalized_key = str(key).strip().casefold().replace("-", "_")
            if normalized_key in PROHIBITED_SESSION_FIELDS:
                prohibited.add(str(key))
            prohibited.update(find_session_fields(nested_value))
    elif isinstance(value, list):
        for nested_value in value:
            prohibited.update(find_session_fields(nested_value))
    return prohibited


def reject_session_fields(source: dict[str, Any]) -> None:
    prohibited = sorted(find_session_fields(source), key=str.casefold)
    if prohibited:
        raise ValueError(
            "snapshot intake must not contain credentials or session data: "
            + ", ".join(prohibited)
        )


def end_of_precision(start: datetime, precision: str) -> datetime:
    normalized = precision.upper()
    if normalized == "DATE":
        return start.replace(hour=23, minute=59, second=59, microsecond=0)
    if normalized == "HOUR":
        return start.replace(microsecond=0) + timedelta(hours=1) - timedelta(seconds=1)
    if normalized == "MINUTE":
        return start.replace(microsecond=0) + timedelta(minutes=1) - timedelta(seconds=1)
    return start.replace(microsecond=0)


def candidate_key(source: dict[str, Any], fallback: str) -> str:
    return str(
        source.get("candidate_key")
        or source.get("material_id")
        or source.get("source_id")
        or source.get("title")
        or fallback
    )


def publication_time_window(source: dict[str, Any]) -> dict[str, Any]:
    raw_window = source.get("publication_time_window")
    if isinstance(raw_window, dict):
        earliest_value = raw_window.get("earliest")
        latest_value = raw_window.get("latest")
        if earliest_value and latest_value:
            earliest = parse_timestamp(str(earliest_value), "publication_time_window.earliest")
            latest = parse_timestamp(str(latest_value), "publication_time_window.latest")
            if earliest > latest:
                raise ValueError("publication_time_window earliest must not be later than latest")
            return {
                "raw_text": str(
                    raw_window.get("raw_text")
                    or source.get("published_at_text")
                    or source.get("published_at")
                    or earliest_value
                ),
                "precision": str(raw_window.get("precision") or source.get("publication_precision") or "SECOND"),
                "earliest": earliest.isoformat(),
                "latest": latest.isoformat(),
                "basis": str(raw_window.get("basis") or source.get("publication_basis") or "intake metadata"),
            }

    published_at = source.get("published_at")
    if not published_at:
        return {
            "raw_text": source.get("published_at_text"),
            "precision": source.get("publication_precision"),
            "earliest": None,
            "latest": None,
            "basis": source.get("publication_basis"),
        }
    precision = str(source.get("publication_precision") or "SECOND")
    earliest = parse_timestamp(str(published_at), "published_at")
    latest = end_of_precision(earliest, precision)
    return {
        "raw_text": str(source.get("published_at_text") or published_at),
        "precision": precision,
        "earliest": earliest.isoformat(),
        "latest": latest.isoformat(),
        "basis": str(source.get("publication_basis") or "intake published_at metadata"),
    }


def displayed_publisher(source: dict[str, Any]) -> str | None:
    value = source.get("displayed_publisher") or source.get("publisher")
    return str(value) if value else None


def acquisition_source(source: dict[str, Any]) -> dict[str, Any] | None:
    raw = source.get("acquisition_source")
    if isinstance(raw, dict) and raw:
        return raw
    source_url = source.get("source_url")
    if source_url:
        return {"type": "url", "url": str(source_url)}
    locator = source.get("locator")
    if isinstance(locator, dict):
        source_page = locator.get("source_page") or locator.get("source_url")
        if source_page:
            return {"type": "url", "url": str(source_page)}
    return None


def material_locator(source: dict[str, Any]) -> dict[str, Any] | None:
    raw = (
        source.get("canonical_material_locator")
        or source.get("canonical_locator")
        or source.get("locator")
    )
    if isinstance(raw, dict) and raw:
        return raw
    if isinstance(raw, str) and raw.strip():
        return {"location": raw.strip()}
    return None


def acquisition_targets(source: dict[str, Any]) -> list[str]:
    raw_targets = (
        source.get("acquisition_targets")
        or source.get("acquisition_target_ids")
        or source.get("target_ids")
    )
    if raw_targets is None and isinstance(source.get("coverage"), dict):
        raw_targets = list(source["coverage"].keys())
    if not isinstance(raw_targets, list):
        return []
    targets: list[str] = []
    for target in raw_targets:
        target_id = target.get("target_id") if isinstance(target, dict) else target
        if target_id:
            targets.append(str(target_id))
    return targets


def source_use_note(source: dict[str, Any]) -> dict[str, Any]:
    note: dict[str, Any] = {}
    for field in (
        "terms_url",
        "known_restrictions",
        "usage_basis",
        "restriction_status",
        "restriction_reason",
        "claimed_original_source",
    ):
        if source.get(field):
            note[field] = source[field]
    return note


def locator_priority(source: dict[str, Any]) -> int:
    source_class = str(source.get("source_class") or source.get("locator_priority") or "").upper()
    if source_class in LOCATOR_PRIORITY:
        return LOCATOR_PRIORITY[source_class]
    publisher = (displayed_publisher(source) or "").casefold()
    source_id = str(source.get("source_id") or "").casefold()
    if "smic" in publisher or "semiconductor manufacturing international" in publisher:
        return LOCATOR_PRIORITY["ISSUER"]
    if any(token in publisher or token in source_id for token in ("sse", "hkex", "exchange", "regulator")):
        return LOCATOR_PRIORITY["REGULATOR_EXCHANGE"]
    if publisher:
        return LOCATOR_PRIORITY["NAMED_INSTITUTION"]
    return LOCATOR_PRIORITY["OTHER"]


def locator_record(source: dict[str, Any]) -> dict[str, Any]:
    locator = material_locator(source)
    if locator is None:
        raise ValueError("source has no material locator")
    record = {
        "locator": locator,
        "acquisition_source": acquisition_source(source),
        "displayed_publisher": displayed_publisher(source),
        "source_id": source.get("source_id"),
        "source_class": source.get("source_class") or source.get("locator_priority") or "OTHER",
        "priority": locator_priority(source),
    }
    if source.get("material_id"):
        record["material_id_hint"] = source["material_id"]
    return record


def missing_minimum_metadata(source: dict[str, Any], window: dict[str, Any]) -> list[str]:
    missing: list[str] = []
    if not acquisition_source(source):
        missing.append("acquisition_source")
    if not displayed_publisher(source):
        missing.append("displayed_publisher")
    if not window.get("earliest") or not window.get("latest") or not window.get("basis"):
        missing.append("publication_time_window")
    if not material_locator(source):
        missing.append("material_locator")
    if not acquisition_targets(source):
        missing.append("acquisition_targets")
    return missing


def holding_candidate_record(
    source: dict[str, Any],
    reason: str,
    missing_metadata: list[str],
    discovered_at: datetime,
) -> dict[str, Any]:
    return {
        "candidate_key": candidate_key(source, f"candidate-{discovered_at.timestamp()}"),
        "title": source.get("title"),
        "discovered_at": discovered_at.isoformat(),
        "known_acquisition_source": acquisition_source(source),
        "known_displayed_publisher": displayed_publisher(source),
        "known_publication_time_window": publication_time_window(source),
        "known_material_locator": material_locator(source),
        "acquisition_targets": acquisition_targets(source),
        "missing_metadata": missing_metadata,
        "holding_reason": reason,
        "user_confirmation_status": source.get("user_confirmation_status", "PENDING_USER_CONFIRMATION"),
    }


def exclusion_record(source: dict[str, Any], reason: str, details: str) -> dict[str, Any]:
    return {
        "material_id_hint": source.get("material_id"),
        "source_id": source.get("source_id"),
        "title": source.get("title"),
        "displayed_publisher": displayed_publisher(source),
        "exclusion_reason": reason,
        "details": details,
        "acquisition_targets": acquisition_targets(source),
    }


def suffix_for_material(source: dict[str, Any]) -> str:
    download_url = source.get("download_url")
    if download_url:
        suffix = Path(urllib.parse.urlparse(str(download_url)).path).suffix.lower()
        if suffix:
            return suffix
    local_path = source.get("local_path")
    if local_path:
        suffix = Path(str(local_path)).suffix.lower()
        if suffix:
            return suffix
    media_type = str(source["media_type"])
    return {
        "application/pdf": ".pdf",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "text/html": ".html",
        "text/plain": ".txt",
    }.get(media_type, ".bin")


def download_public_material(download_url: str, destination: Path) -> None:
    parsed_url = urllib.parse.urlparse(download_url)
    if parsed_url.scheme not in {"http", "https"}:
        raise ValueError("download_url must use http or https")
    if parsed_url.username or parsed_url.password:
        raise ValueError("download_url must not contain credentials")
    request = urllib.request.Request(
        download_url,
        headers={"User-Agent": "TradingAgent-mvp-snapshot-builder/2.0"},
    )
    partial_path = destination.with_suffix(destination.suffix + ".partial")
    downloaded = 0
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            final_url = urllib.parse.urlparse(response.geturl())
            if final_url.scheme not in {"http", "https"}:
                raise ValueError("download redirected to a non-http(s) URL")
            with partial_path.open("wb") as output:
                while chunk := response.read(1024 * 1024):
                    downloaded += len(chunk)
                    if downloaded > MAX_DOWNLOAD_BYTES:
                        raise ValueError("download exceeds the 100 MiB snapshot limit")
                    output.write(chunk)
        partial_path.replace(destination)
    finally:
        if partial_path.exists():
            partial_path.unlink()


def acquire_bytes(source: dict[str, Any], intake_file: Path, destination: Path) -> str:
    has_local_path = bool(source.get("local_path"))
    has_download_url = bool(source.get("download_url"))
    if has_local_path == has_download_url:
        raise ValueError(
            f"material {source.get('material_id')} must have exactly one of local_path or download_url"
        )
    if has_local_path:
        local_path = (intake_file.parent / str(source["local_path"])).resolve()
        if not local_path.is_file():
            raise ValueError(f"material file does not exist: {local_path}")
        shutil.copyfile(local_path, destination)
        return "LOCAL_PREPARED"
    download_public_material(str(source["download_url"]), destination)
    return "DIRECT_PUBLIC_DOWNLOAD"


def extract_text(path: Path, material_id: str) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8")
    return {
        "material_id": material_id,
        "parser": "plain-text-v1",
        "lines": [
            {"line_number": number, "text": line}
            for number, line in enumerate(text.splitlines(), start=1)
        ],
    }


def extract_pdf(path: Path, material_id: str) -> dict[str, Any]:
    parsed_pages: list[dict[str, Any]] = []
    extracted_character_count = 0
    with pdfplumber.open(path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            extracted_character_count += len(re.sub(r"\s+", "", text))
            text_lines: list[dict[str, Any]] = []
            for line in page.extract_text_lines(return_chars=False):
                text_lines.append(
                    {
                        "text": line["text"],
                        "bbox": [line["x0"], line["top"], line["x1"], line["bottom"]],
                    }
                )
            tables: list[dict[str, Any]] = []
            for table_number, table in enumerate(page.find_tables(), start=1):
                tables.append(
                    {
                        "table_number": table_number,
                        "bbox": list(table.bbox),
                        "rows": table.extract(),
                    }
                )
            section_candidates = [
                line["text"].strip()
                for line in text_lines
                if line["text"].strip().isupper()
                and 2 <= len(line["text"].strip()) <= 160
            ]
            footnote_candidates = [
                line
                for line in text_lines
                if line["bbox"][1] >= page.height * 0.65
                or re.match(r"^\s*(?:\d+|\*+)\s+", line["text"])
            ]
            parsed_pages.append(
                {
                    "page_number": page_number,
                    "width": page.width,
                    "height": page.height,
                    "text": text,
                    "text_lines": text_lines,
                    "tables": tables,
                    "section_candidates": section_candidates,
                    "footnote_candidates": footnote_candidates,
                }
            )
    reliable_text_layer = bool(parsed_pages) and extracted_character_count >= (
        20 * len(parsed_pages)
    )
    return {
        "material_id": material_id,
        "parser": "pdfplumber-v1",
        "page_count": len(parsed_pages),
        "reliable_text_layer": reliable_text_layer,
        "pages": parsed_pages,
    }


def normalize_html_text(parts: list[str]) -> str:
    return re.sub(r"\s+", " ", "".join(parts)).strip()


class ResearchHtmlParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.headings: list[dict[str, Any]] = []
        self.paragraphs: list[dict[str, Any]] = []
        self.tables: list[dict[str, Any]] = []
        self._block_tag: str | None = None
        self._block_parts: list[str] = []
        self._table_rows: list[list[str]] | None = None
        self._row: list[str] | None = None
        self._cell_parts: list[str] | None = None

    def handle_starttag(
        self, tag: str, attrs: list[tuple[str, str | None]]
    ) -> None:
        del attrs
        if tag in {"h1", "h2", "h3", "h4", "h5", "h6", "p"}:
            self._block_tag = tag
            self._block_parts = []
        elif tag == "table":
            self._table_rows = []
        elif tag == "tr" and self._table_rows is not None:
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            self._cell_parts = []

    def handle_data(self, data: str) -> None:
        if self._block_tag is not None:
            self._block_parts.append(data)
        if self._cell_parts is not None:
            self._cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self._cell_parts is not None:
            if self._row is not None:
                self._row.append(normalize_html_text(self._cell_parts))
            self._cell_parts = None
        elif tag == "tr" and self._row is not None:
            if self._table_rows is not None:
                self._table_rows.append(self._row)
            self._row = None
        elif tag == "table" and self._table_rows is not None:
            table_number = len(self.tables) + 1
            self.tables.append(
                {
                    "table_number": table_number,
                    "locator": f"table[{table_number}]",
                    "rows": self._table_rows,
                }
            )
            self._table_rows = None
        elif tag == self._block_tag:
            text = normalize_html_text(self._block_parts)
            if text:
                if tag == "p":
                    index = len(self.paragraphs) + 1
                    self.paragraphs.append(
                        {"locator": f"paragraph:p[{index}]", "text": text}
                    )
                else:
                    level = int(tag[1])
                    index = sum(
                        heading["level"] == level for heading in self.headings
                    ) + 1
                    self.headings.append(
                        {
                            "level": level,
                            "locator": f"heading:{tag}[{index}]",
                            "text": text,
                        }
                    )
            self._block_tag = None
            self._block_parts = []


def extract_html(path: Path, material_id: str) -> dict[str, Any]:
    parser = ResearchHtmlParser()
    parser.feed(path.read_text(encoding="utf-8"))
    parser.close()
    return {
        "material_id": material_id,
        "parser": "html-parser-v1",
        "headings": parser.headings,
        "paragraphs": parser.paragraphs,
        "tables": parser.tables,
    }


def normalize_spreadsheet_value(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def extract_xlsx(path: Path, material_id: str) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            rows: list[dict[str, Any]] = []
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = [normalize_spreadsheet_value(value) for value in row]
                if any(value is not None for value in values):
                    rows.append({"row_number": row_number, "values": values})
            sheets.append({"name": worksheet.title, "rows": rows})
    finally:
        workbook.close()
    return {
        "material_id": material_id,
        "parser": "openpyxl-v1",
        "sheets": sheets,
    }


def extract_material(path: Path, material_id: str, media_type: str) -> dict[str, Any]:
    if media_type == "application/pdf" or path.suffix.lower() == ".pdf":
        return extract_pdf(path, material_id)
    if media_type == "text/html" or path.suffix.lower() in {".html", ".htm"}:
        return extract_html(path, material_id)
    if (
        media_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        or path.suffix.lower() == ".xlsx"
    ):
        return extract_xlsx(path, material_id)
    if media_type.startswith("text/") or path.suffix.lower() in {".txt", ".md"}:
        return extract_text(path, material_id)
    raise ValueError(f"unsupported media type: {media_type}")


def parse_frozen_material(
    frozen_path: Path,
    parsed_dir: Path,
    material_id: str,
    media_type: str,
) -> tuple[str, str | None, dict[str, Any] | None]:
    parsed_path = parsed_dir / f"{material_id}.json"
    try:
        parsed_material = extract_material(frozen_path, material_id, media_type)
    except ValueError as error:
        if "unsupported media type" not in str(error):
            raise
        return (
            "UNSUPPORTED",
            None,
            {
                "gap_kind": "MATERIAL_UNPARSEABLE",
                "question": f"Current parser cannot process material {material_id}: {media_type}",
                "current_handling": "Freeze the raw file only; do not generate chunks, facts, or evidence.",
                "required_evidence": "A text-layer PDF, HTML, text, or spreadsheet representation of the same material.",
            },
        )
    except Exception as error:  # noqa: BLE001 - parser failure should not drop frozen bytes.
        write_json(
            parsed_path,
            {
                "material_id": material_id,
                "parser": "parser-error-v1",
                "error": str(error),
            },
        )
        return (
            "PARTIAL",
            parsed_path.name,
            {
                "gap_kind": "MATERIAL_UNPARSEABLE",
                "question": f"Parser failed before reliable extraction for material {material_id}.",
                "current_handling": "Freeze the raw file and preserve parser diagnostics only.",
                "required_evidence": "A representation that the current Demo parser can read without errors.",
            },
        )

    write_json(parsed_path, parsed_material)
    if media_type == "application/pdf" and not parsed_material.get("reliable_text_layer"):
        return (
            "UNSUPPORTED",
            parsed_path.name,
            {
                "gap_kind": "MATERIAL_UNPARSEABLE",
                "question": f"PDF material {material_id} does not expose a reliable text layer.",
                "current_handling": "Do not OCR or read values from page images; keep raw bytes and parser diagnostics.",
                "required_evidence": "A text-layer PDF or official text/table version of the same material.",
            },
        )
    return "PARSED", parsed_path.name, None


def acquisition_gap(
    gap_id: str,
    gap_kind: str,
    question: str,
    impact_objects: list[str],
    current_handling: str,
    required_evidence: str,
    priority: str = "P1",
) -> dict[str, Any]:
    return {
        "gap_id": gap_id,
        "origin_stage": "acquire-research-materials",
        "gap_kind": gap_kind,
        "question": question,
        "impact_objects": impact_objects,
        "current_handling": current_handling,
        "confirmation_owner": "user/data_preparer",
        "required_evidence": required_evidence,
        "priority": priority,
        "status": "OPEN",
    }


def canonical_snapshot_id(
    case: dict[str, Any],
    materials: list[dict[str, Any]],
    intake: dict[str, Any],
    test_fixture: bool,
) -> str:
    volatile_material_fields = {"acquired_at", "frozen_path", "parsed_path"}
    identity = {
        "case": {
            field: case.get(field)
            for field in (
                "case_origin",
                "company",
                "security",
                "as_of",
                "timezone",
                "required_coverage",
                "source_use_assumption",
                "rule_versions",
            )
        },
        "parent_snapshot_id": intake.get("parent_snapshot_id"),
        "acquisition_targets": intake.get("acquisition_targets", []),
        "search_saturation": intake.get("search_saturation", {}),
        "materials": [
            {
                key: value
                for key, value in material.items()
                if key not in volatile_material_fields
            }
            for material in sorted(materials, key=lambda item: item["material_id"])
        ],
    }
    digest = hashlib.sha256(
        json.dumps(identity, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()[:16]
    return f"{'fixture' if test_fixture else 'smic'}-{digest}"


def material_base_record(
    source: dict[str, Any],
    window: dict[str, Any],
    created_at: datetime,
    content_hash: str,
    frozen_path: Path,
    output_dir: Path,
    acquisition_method: str,
    parse_status: str,
    parsed_path_name: str | None,
) -> dict[str, Any]:
    record: dict[str, Any] = {
        "material_id": str(source["material_id"]),
        "source_id": str(source["source_id"]),
        "title": str(source["title"]),
        "displayed_publisher": displayed_publisher(source),
        "publication_time": window,
        "acquired_at": created_at.isoformat(),
        "acquisition_source": acquisition_source(source),
        "canonical_material_locator": material_locator(source),
        "alternate_material_locators": [],
        "media_type": str(source["media_type"]),
        "content_hash": content_hash,
        "frozen_path": frozen_path.relative_to(output_dir).as_posix(),
        "acquisition_targets": acquisition_targets(source),
        "acquisition_status": ACQUISITION_STATUS,
        "as_of_eligible": True,
        "as_of_reason": "publication_time_window.latest is not later than case as_of",
        "parse_status": parse_status,
        "acquisition_method": acquisition_method,
        "source_use_note": source_use_note(source),
    }
    if parsed_path_name is not None:
        record["parsed_path"] = f"parsed/{parsed_path_name}"
    return record


def merge_duplicate_locator(existing: dict[str, Any], source: dict[str, Any]) -> None:
    incoming = locator_record(source)
    current = {
        "locator": existing["canonical_material_locator"],
        "acquisition_source": existing["acquisition_source"],
        "displayed_publisher": existing["displayed_publisher"],
        "source_id": existing["source_id"],
        "source_class": existing.get("canonical_locator_source_class", "OTHER"),
        "priority": existing.get("canonical_locator_priority", LOCATOR_PRIORITY["OTHER"]),
    }
    alternate_records = [current, incoming]
    alternate_records.extend(existing.get("alternate_material_locators", []))
    best = min(alternate_records, key=lambda item: (int(item.get("priority", 99)), str(item.get("source_id"))))
    alternates = [
        item
        for item in alternate_records
        if json.dumps(item.get("locator"), ensure_ascii=False, sort_keys=True)
        != json.dumps(best.get("locator"), ensure_ascii=False, sort_keys=True)
    ]
    existing["canonical_material_locator"] = best["locator"]
    existing["acquisition_source"] = best.get("acquisition_source")
    existing["displayed_publisher"] = best.get("displayed_publisher")
    existing["source_id"] = best.get("source_id")
    existing["canonical_locator_source_class"] = best.get("source_class")
    existing["canonical_locator_priority"] = best.get("priority")
    existing["alternate_material_locators"] = dedupe_locators(alternates)
    existing["acquisition_targets"] = sorted(
        set(existing.get("acquisition_targets", [])) | set(acquisition_targets(source))
    )


def dedupe_locators(locators: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    deduped: list[dict[str, Any]] = []
    for locator in locators:
        key = json.dumps(
            {
                "locator": locator.get("locator"),
                "acquisition_source": locator.get("acquisition_source"),
                "displayed_publisher": locator.get("displayed_publisher"),
                "source_id": locator.get("source_id"),
            },
            ensure_ascii=False,
            sort_keys=True,
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(locator)
    return deduped


def copy_snapshot_inputs(case_file: Path, intake_file: Path, output_dir: Path) -> list[str]:
    copied: list[str] = []
    for source_path, destination_name in (
        (case_file, "case.yaml"),
        (intake_file, "snapshot-inputs.yaml"),
    ):
        destination = output_dir / destination_name
        if source_path.resolve() != destination.resolve():
            shutil.copyfile(source_path, destination)
        copied.append(destination_name)
    return copied


def build_target_summary(
    case: dict[str, Any],
    intake: dict[str, Any],
    material_records: list[dict[str, Any]],
    holding_records: list[dict[str, Any]],
    gaps: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    raw_targets = intake.get("acquisition_targets")
    if not isinstance(raw_targets, list) or not raw_targets:
        raw_targets = [
            {"target_id": target_id, "label": target_id}
            for target_id in case.get("required_coverage", [])
        ]
    gap_by_target: dict[str, list[str]] = {}
    for gap in gaps:
        for impact in gap.get("impact_objects", []):
            gap_by_target.setdefault(str(impact), []).append(str(gap["gap_id"]))
    summaries: list[dict[str, Any]] = []
    for raw_target in raw_targets:
        if isinstance(raw_target, dict):
            target_id = str(raw_target["target_id"])
            label = raw_target.get("label") or target_id
            channels = raw_target.get("search_channels", [])
            status = str(raw_target.get("search_status") or "IN_PROGRESS")
        else:
            target_id = str(raw_target)
            label = target_id
            channels = []
            status = "IN_PROGRESS"
        if status not in ALLOWED_SEARCH_STATUSES:
            raise ValueError(f"invalid search_status for target {target_id}: {status}")
        material_ids = [
            record["material_id"]
            for record in material_records
            if target_id in record.get("acquisition_targets", [])
        ]
        candidate_count = sum(
            target_id in record.get("acquisition_targets", []) for record in holding_records
        )
        summaries.append(
            {
                "target_id": target_id,
                "label": label,
                "search_channels": channels,
                "material_ids": material_ids,
                "candidate_holding_count": candidate_count,
                "gap_ids": sorted(set(gap_by_target.get(target_id, []))),
                "search_status": status,
            }
        )
    return summaries


def snapshot_build(args: argparse.Namespace) -> int:
    case_file = Path(args.case_file).resolve()
    intake_file = Path(args.intake_file).resolve()
    output_dir = Path(args.output_dir).resolve()
    holding_dir = (
        Path(args.holding_dir).resolve()
        if args.holding_dir
        else output_dir.parent / f"{output_dir.name}-candidate-holding"
    )
    case = yaml.safe_load(case_file.read_text(encoding="utf-8"))
    intake = yaml.safe_load(intake_file.read_text(encoding="utf-8"))
    if not isinstance(case, dict) or not isinstance(intake, dict):
        raise ValueError("case and intake files must contain YAML mappings")
    if not case.get("as_of"):
        raise ValueError("case.yaml is missing as_of")
    as_of = parse_timestamp(str(case["as_of"]), "as_of")
    created_at = parse_timestamp(args.created_at, "created_at")

    if (output_dir / "snapshot-manifest.yaml").exists():
        raise ValueError(
            "snapshot directory is already frozen; use a new output directory for changed materials"
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    raw_dir = output_dir / "raw"
    parsed_dir = output_dir / "parsed"
    raw_dir.mkdir(exist_ok=True)
    parsed_dir.mkdir(exist_ok=True)

    copied_input_files = copy_snapshot_inputs(case_file, intake_file, output_dir)
    material_records: list[dict[str, Any]] = []
    records_by_hash: dict[str, dict[str, Any]] = {}
    holding_records: list[dict[str, Any]] = []
    excluded_records: list[dict[str, Any]] = []
    gaps: list[dict[str, Any]] = []
    seen_material_ids: set[str] = set()
    gap_counter = 1

    for candidate in intake.get("candidate_holding_area", []):
        if not isinstance(candidate, dict):
            raise ValueError("each candidate_holding_area entry must be a YAML mapping")
        reject_session_fields(candidate)
        window = publication_time_window(candidate)
        missing = missing_minimum_metadata(candidate, window)
        holding_records.append(
            holding_candidate_record(candidate, "INTAKE_DECLARED_CANDIDATE", missing, created_at)
        )
        held_candidate_key = candidate_key(candidate, "candidate")
        gaps.append(
            acquisition_gap(
                f"GAP_ACQUIRE_{gap_counter:03d}",
                "CANDIDATE_HOLDING_PENDING_METADATA",
                f"Candidate {held_candidate_key} remains outside every snapshot pending metadata or user confirmation.",
                acquisition_targets(candidate) or [held_candidate_key],
                "Do not assign a material_id or use the candidate for RAG, facts, analysis, or report output.",
                "Complete acquisition minimum metadata and user confirmation before a future snapshot build.",
            )
        )
        gap_counter += 1

    for source in intake.get("materials", []):
        if not isinstance(source, dict):
            raise ValueError("each material entry must be a YAML mapping")
        reject_session_fields(source)
        require_fields(source, ("material_id", "source_id", "title", "media_type"))
        material_id = str(source["material_id"])
        validate_material_id(material_id, seen_material_ids)
        window = publication_time_window(source)
        missing_metadata = missing_minimum_metadata(source, window)
        if missing_metadata:
            holding_records.append(
                holding_candidate_record(source, "MISSING_MINIMUM_METADATA", missing_metadata, created_at)
            )
            gaps.append(
                acquisition_gap(
                    f"GAP_ACQUIRE_{gap_counter:03d}",
                    "CANDIDATE_HOLDING_PENDING_METADATA",
                    f"Candidate {material_id} is missing acquisition minimum metadata: {', '.join(missing_metadata)}.",
                    acquisition_targets(source) or [material_id],
                    "Keep the candidate outside every snapshot until metadata is supplied and confirmed.",
                    "Traceable acquisition source, displayed publisher, publication time window, material locator, and target mapping.",
                )
            )
            gap_counter += 1
            continue

        latest = parse_timestamp(str(window["latest"]), "publication_time_window.latest")
        earliest = parse_timestamp(str(window["earliest"]), "publication_time_window.earliest")
        if earliest > as_of:
            excluded_records.append(
                exclusion_record(
                    source,
                    "AS_OF_EXCEEDED",
                    "publication_time_window.earliest is later than case as_of",
                )
            )
            continue
        if latest > as_of:
            holding_records.append(
                holding_candidate_record(
                    source,
                    "PUBLICATION_WINDOW_CROSSES_AS_OF",
                    ["publication_time_window_after_as_of_boundary"],
                    created_at,
                )
            )
            gaps.append(
                acquisition_gap(
                    f"GAP_ACQUIRE_{gap_counter:03d}",
                    "CANDIDATE_HOLDING_PENDING_METADATA",
                    f"Candidate {material_id} publication window crosses case as_of.",
                    acquisition_targets(source),
                    "Keep the candidate outside every snapshot until the publication boundary is confirmed.",
                    "Publication evidence proving the latest possible public time is not later than case as_of.",
                )
            )
            gap_counter += 1
            continue

        temporary_path = raw_dir / f"{material_id}.download{suffix_for_material(source)}"
        try:
            acquisition_method = acquire_bytes(source, intake_file, temporary_path)
        except ValueError as error:
            gaps.append(
                acquisition_gap(
                    f"GAP_ACQUIRE_{gap_counter:03d}",
                    "MATERIAL_UNAVAILABLE",
                    f"Material {material_id} could not be frozen: {error}",
                    acquisition_targets(source),
                    "Do not include the material in this snapshot.",
                    "A public download or user-prepared local file with reproducible bytes.",
                )
            )
            gap_counter += 1
            continue

        content_hash = sha256_file(temporary_path)
        if content_hash in records_by_hash:
            temporary_path.unlink()
            merge_duplicate_locator(records_by_hash[content_hash], source)
            continue

        frozen_path = raw_dir / f"{material_id}{suffix_for_material(source)}"
        temporary_path.replace(frozen_path)
        parse_status, parsed_path_name, parse_gap = parse_frozen_material(
            frozen_path,
            parsed_dir,
            material_id,
            str(source["media_type"]),
        )
        record = material_base_record(
            source,
            window,
            created_at,
            content_hash,
            frozen_path,
            output_dir,
            acquisition_method,
            parse_status,
            parsed_path_name,
        )
        canonical = locator_record(source)
        record["canonical_locator_source_class"] = canonical["source_class"]
        record["canonical_locator_priority"] = canonical["priority"]
        material_records.append(record)
        records_by_hash[content_hash] = record
        if parse_gap:
            gaps.append(
                acquisition_gap(
                    f"GAP_ACQUIRE_{gap_counter:03d}",
                    str(parse_gap["gap_kind"]),
                    str(parse_gap["question"]),
                    [material_id, *acquisition_targets(source)],
                    str(parse_gap["current_handling"]),
                    str(parse_gap["required_evidence"]),
                )
            )
            gap_counter += 1

    for raw_gap in intake.get("acquisition_gaps", []):
        if not isinstance(raw_gap, dict):
            raise ValueError("each acquisition_gaps entry must be a YAML mapping")
        gap_kind = str(raw_gap.get("gap_kind") or raw_gap.get("kind") or "SEARCH_NOT_FOUND")
        targets = [str(item) for item in raw_gap.get("acquisition_targets", [])]
        gaps.append(
            acquisition_gap(
                f"GAP_ACQUIRE_{gap_counter:03d}",
                gap_kind,
                str(raw_gap["question"]),
                targets or [str(raw_gap.get("target_id") or "acquisition")],
                str(raw_gap.get("current_handling") or "Record the acquisition gap; do not infer facts."),
                str(raw_gap.get("required_evidence") or "Additional traceable material or user confirmation."),
                str(raw_gap.get("priority") or "P1"),
            )
        )
        gap_counter += 1

    materials_path = output_dir / "materials.jsonl"
    materials_path.write_text(
        "".join(
            json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n"
            for record in sorted(material_records, key=lambda item: item["material_id"])
        ),
        encoding="utf-8",
    )

    test_fixture = case.get("case_origin") == "fixture"
    snapshot_id = canonical_snapshot_id(case, material_records, intake, test_fixture)
    gaps_path = output_dir / "gaps.yaml"
    write_yaml(gaps_path, {"snapshot_id": snapshot_id, "gaps": gaps})

    holding_summary: dict[str, Any] = {"count": len(holding_records)}
    if holding_records:
        holding_dir.mkdir(parents=True, exist_ok=True)
        holding_path = holding_dir / "candidates.yaml"
        write_yaml(
            holding_path,
            {
                "related_snapshot_build": snapshot_id,
                "snapshot_membership": "OUTSIDE_ALL_SNAPSHOTS",
                "candidates": holding_records,
            },
        )
        holding_summary.update(
            {
                "path": holding_path.relative_to(output_dir.parent).as_posix(),
                "hash": sha256_file(holding_path),
            }
        )

    target_summary = build_target_summary(
        case, intake, material_records, holding_records, gaps
    )
    search_saturation = intake.get("search_saturation", {})
    if not isinstance(search_saturation, dict):
        raise ValueError("search_saturation must be a YAML mapping when present")
    global_search_status = str(search_saturation.get("status") or "IN_PROGRESS")
    if global_search_status not in ALLOWED_SEARCH_STATUSES:
        raise ValueError(f"invalid search_saturation status: {global_search_status}")

    manifest_files = [
        {"path": relative_path, "hash": sha256_file(output_dir / relative_path)}
        for relative_path in copied_input_files
    ]
    manifest_files.extend(
        [
            {"path": "materials.jsonl", "hash": sha256_file(materials_path)},
            {"path": "gaps.yaml", "hash": sha256_file(gaps_path)},
            *[
                {"path": record["frozen_path"], "hash": record["content_hash"]}
                for record in material_records
            ],
            *[
                {
                    "path": record["parsed_path"],
                    "hash": sha256_file(output_dir / record["parsed_path"]),
                }
                for record in material_records
                if "parsed_path" in record
            ],
        ]
    )
    manifest = {
        "snapshot_id": snapshot_id,
        "parent_snapshot_id": intake.get("parent_snapshot_id"),
        "created_at": created_at.isoformat(),
        "as_of": case["as_of"],
        "build_mode": "SNAPSHOT_BUILD",
        "test_fixture": test_fixture,
        "usage_assumption": "personal, non-commercial, local, not for redistribution or model training",
        "statistics": {
            "unique_materials": len(material_records),
            "material_locators": sum(
                1 + len(record.get("alternate_material_locators", []))
                for record in material_records
            ),
            "excluded_after_as_of": sum(
                record["exclusion_reason"] == "AS_OF_EXCEEDED"
                for record in excluded_records
            ),
            "candidate_holding": len(holding_records),
            "parse_unsupported": sum(
                record["parse_status"] == "UNSUPPORTED" for record in material_records
            ),
            "parse_partial": sum(
                record["parse_status"] == "PARTIAL" for record in material_records
            ),
            "acquisition_gaps": len(gaps),
        },
        "acquisition_status": ACQUISITION_STATUS,
        "candidate_holding_area": holding_summary,
        "excluded_materials": excluded_records,
        "search_saturation": {**search_saturation, "status": global_search_status},
        "acquisition_targets": target_summary,
        "files": manifest_files,
    }
    write_yaml(output_dir / "snapshot-manifest.yaml", manifest)
    print(snapshot_id)
    return 0


def validate_snapshot_file_paths(snapshot_dir: Path, manifest: dict[str, Any]) -> None:
    for file_entry in manifest.get("files", []):
        relative_path = Path(str(file_entry["path"]))
        if relative_path.is_absolute():
            raise ValueError(f"snapshot file path must be relative: {relative_path}")
        frozen_file = (snapshot_dir / relative_path).resolve()
        if not frozen_file.is_relative_to(snapshot_dir):
            raise ValueError(f"snapshot file path escapes snapshot directory: {relative_path}")
        if not frozen_file.is_file():
            raise ValueError(f"snapshot file is missing: {relative_path.as_posix()}")
        actual_hash = sha256_file(frozen_file)
        if actual_hash != file_entry["hash"]:
            raise ValueError(f"snapshot file hash mismatch: {relative_path.as_posix()}")


def validate_material_record(material: dict[str, Any], as_of: datetime, line_number: int, snapshot_dir: Path) -> None:
    required_fields = (
        "material_id",
        "source_id",
        "title",
        "displayed_publisher",
        "publication_time",
        "acquisition_source",
        "canonical_material_locator",
        "media_type",
        "content_hash",
        "frozen_path",
        "acquisition_targets",
        "acquisition_status",
        "as_of_eligible",
        "parse_status",
    )
    missing = [field for field in required_fields if not material.get(field)]
    if missing:
        raise ValueError(f"snapshot material is missing required fields at line {line_number}: {', '.join(missing)}")
    if material["acquisition_status"] != ACQUISITION_STATUS:
        raise ValueError(f"snapshot material is not ACQUIRED_UNASSESSED at line {line_number}")
    if material["parse_status"] not in ALLOWED_PARSE_STATUSES:
        raise ValueError(f"snapshot material has invalid parse_status at line {line_number}")
    window = material["publication_time"]
    if not isinstance(window, dict):
        raise ValueError(f"snapshot material has invalid publication_time at line {line_number}")
    missing_window_fields = [
        field
        for field in ("raw_text", "precision", "earliest", "latest", "basis")
        if not window.get(field)
    ]
    if missing_window_fields:
        raise ValueError(
            "snapshot material publication_time is missing required fields "
            f"at line {line_number}: {', '.join(missing_window_fields)}"
        )
    parse_timestamp(str(window["earliest"]), "publication_time.earliest")
    latest = parse_timestamp(str(window["latest"]), "publication_time.latest")
    if latest > as_of or not material["as_of_eligible"]:
        raise ValueError(f"included material exceeds as_of at line {line_number}")
    targets = material["acquisition_targets"]
    if not isinstance(targets, list) or not targets:
        raise ValueError(f"snapshot material has no acquisition target at line {line_number}")
    frozen_path = Path(str(material["frozen_path"]))
    if frozen_path.is_absolute():
        raise ValueError(f"material frozen_path must be relative at line {line_number}")
    frozen_file = (snapshot_dir / frozen_path).resolve()
    if not frozen_file.is_relative_to(snapshot_dir):
        raise ValueError(f"material frozen_path escapes snapshot directory at line {line_number}")
    if sha256_file(frozen_file) != material["content_hash"]:
        raise ValueError(f"material content hash mismatch at line {line_number}")


def read_materials(snapshot_dir: Path, as_of: datetime) -> list[dict[str, Any]]:
    materials_path = snapshot_dir / "materials.jsonl"
    material_records: list[dict[str, Any]] = []
    seen_hashes: set[str] = set()
    for line_number, line in enumerate(
        materials_path.read_text(encoding="utf-8").splitlines(), start=1
    ):
        if not line.strip():
            continue
        material = json.loads(line)
        if not isinstance(material, dict):
            raise ValueError(f"snapshot material is not an object at line {line_number}")
        reject_session_fields(material)
        validate_material_record(material, as_of, line_number, snapshot_dir)
        if material["content_hash"] in seen_hashes:
            raise ValueError(f"duplicate content_hash in materials.jsonl at line {line_number}")
        seen_hashes.add(material["content_hash"])
        material_records.append(material)
    return material_records


def demo_run(args: argparse.Namespace) -> int:
    case_file = Path(args.case_file).resolve()
    snapshot_dir = Path(args.snapshot_dir).resolve()
    case = yaml.safe_load(case_file.read_text(encoding="utf-8"))
    manifest = yaml.safe_load(
        (snapshot_dir / "snapshot-manifest.yaml").read_text(encoding="utf-8")
    )
    intake = yaml.safe_load(
        (snapshot_dir / "snapshot-inputs.yaml").read_text(encoding="utf-8")
    )
    if not isinstance(case, dict) or not isinstance(manifest, dict) or not isinstance(intake, dict):
        raise ValueError("case, snapshot manifest, and snapshot inputs must contain YAML mappings")
    if manifest.get("as_of") != case.get("as_of"):
        raise ValueError("snapshot as_of does not match case as_of")
    if manifest.get("build_mode") != "SNAPSHOT_BUILD":
        raise ValueError("snapshot manifest build_mode must be SNAPSHOT_BUILD")
    as_of = parse_timestamp(str(case["as_of"]), "as_of")

    validate_snapshot_file_paths(snapshot_dir, manifest)
    material_records = read_materials(snapshot_dir, as_of)
    expected_snapshot_id = canonical_snapshot_id(
        case, material_records, intake, case.get("case_origin") == "fixture"
    )
    if manifest.get("snapshot_id") != expected_snapshot_id:
        raise ValueError("snapshot_id does not match the frozen case, intake, and material identity")

    print(manifest["snapshot_id"])
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Freeze the single-stock Demo research snapshot")
    subparsers = parser.add_subparsers(dest="mode", required=True)
    build = subparsers.add_parser("SNAPSHOT_BUILD")
    build.add_argument("--case-file", required=True)
    build.add_argument("--intake-file", required=True)
    build.add_argument("--output-dir", required=True)
    build.add_argument("--created-at", required=True)
    build.add_argument("--holding-dir")
    demo = subparsers.add_parser("DEMO_RUN")
    demo.add_argument("--case-file", required=True)
    demo.add_argument("--snapshot-dir", required=True)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        if args.mode == "SNAPSHOT_BUILD":
            return snapshot_build(args)
        if args.mode == "DEMO_RUN":
            return demo_run(args)
        raise ValueError(f"unsupported mode: {args.mode}")
    except (OSError, ValueError, yaml.YAMLError, json.JSONDecodeError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
